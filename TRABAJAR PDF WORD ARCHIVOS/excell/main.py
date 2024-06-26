import openpyxl
from openpyxl import Workbook
path = 'sueltos.xlsx'

wb_obj = openpyxl.load_workbook(path)

organizado = openpyxl.Workbook()
sheet = organizado.active

sheet_obj = wb_obj.active

row = sheet_obj.max_row
column = sheet_obj.max_column


for fila in range(2, row):
    nombre = sheet_obj.cell(row=fila, column = 5)
    nombre_str = nombre.value
    if nombre_str.__contains__('First name: '):# or nombre_str.__contains__('Full name: '):
        nombre_str = nombre_str[:nombre_str.index('First name:')]
    #print(nombre.value)

    llamadas = sheet_obj.cell(row=fila, column = 6)
    llamadas.value = llamadas.value.replace('incoming','entrante')
    llamadas.value = llamadas.value.replace('outgoing','saliente')
    llamadas.value = llamadas.value.replace('missed','perdida')
    llamadas.value = llamadas.value.replace('undelivered','no entregado')
    #print('\t' + llamadas.value)

    mensajes = sheet_obj.cell(row=fila, column = 7)
    mensajes.value = mensajes.value.replace('incoming','entrante')
    mensajes.value = mensajes.value.replace('outgoing','saliente')
    mensajes.value = mensajes.value.replace('missed','perdida')
    mensajes.value = mensajes.value.replace('undelivered','no entregado')
    #print('\t' + mensajes.value)

    telefono = sheet_obj.cell(row=fila, column=8)
    telefono_str = telefono.value


    if telefono_str.__contains__('Móvil'):
        telefono_str = telefono_str[telefono_str.rindex('Móvil: '):]
        telefono_str = telefono_str[7:]
    elif telefono_str.__contains__('Phone number'):
        telefono_str = telefono_str[telefono_str.rindex('number: '):]
        telefono_str = telefono_str[7:]
    elif telefono_str.__contains__('Mobile: '):
        telefono_str = telefono_str[telefono_str.rindex('Mobile: '):]
        telefono_str = telefono_str[7:]
    else:
        telefono_str = 'N/A'
    #print(telefono_str)
    sheet.cell(row = fila-1, column = 1).value = nombre_str
    sheet.cell(row = fila-1, column = 2).value = telefono_str
    sheet.cell(row = fila-1, column = 3).value = llamadas.value
    sheet.cell(row = fila-1, column = 4).value = mensajes.value

organizado.save('organizados.xlsx')